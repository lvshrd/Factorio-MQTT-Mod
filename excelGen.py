# This script generates an Excel file for importing data models into the supOS platform.
# The excel should include a `folder` sheet and the corresponding DB sheet, eg `relation`, `TimeSeries`
# All the data model structure is needed for each single topic, eg for 'Factorio/Sandbox/Smelting/Line2530/furnace2415/elec'
# It needs blank rows like `Factorio/Sandbox/`, `Factorio/Sandbox/Smelting/`, `Factorio/Sandbox/Smelting/Line2530/`, etc.
# The main components of the script are:
# 1. ExcelGenerator class:
#     - Initializes with a template Excel file and an output filename.
#     - Analyzes topics and payloads to categorize them into folder hierarchy, relation data, and time series data.
#     - Saves the categorized data into an Excel file with three sheets: Folder, Relation, and TimeSeries.
# 2. add_topic method:
#     - Analyzes a given topic and its payload.
#     - Categorizes the topic into folder hierarchy, relation data, or time series data based on predefined rules.
# 3. save_excel method:
#     - Copies the template Excel file to the output filename.
#     - Writes the categorized data into the respective sheets in the Excel file.
# 4. main function:
#     - Reads a log file containing topics and payloads.
#     - Uses the ExcelGenerator to process each topic and payload.
#     - Saves the processed data into an Excel file.

    
from openpyxl import load_workbook
import json
import os
from config import LOG_FILE
import shutil
from datetime import datetime

class ExcelGenerator:
    def __init__(self, template_filename, output_filename):
        self.template_filename = template_filename
        self.output_filename = output_filename
        self.folder_data = set()       # Set to avoid duplicate
        self.relation_data = {}        # Relation Sheet: {topic: fields}
        self.timeseries_data = {}      # TimeSeries Sheet: {topic: fields}

    def add_topic(self, full_topic, payload):
        """
        Analyze Topic and fill into three sheets
        :param full_topic: Full Topic path, eg Factorio/demo/test/Power/Line76/generator1578/elec
        :param payload: The JSON data corresponding to the Topic
        """
        # 1. 处理 Folder Sheet（目录层级）
        parts = full_topic.split('/')
        current_path = ""
        for part in parts[:-1]:  # 排除最后一个节点（如 elec）
            current_path += part + "/"
            # 去掉只有第一个层级的topic(eg: Factorio/)(supOS平台导入限制)
            if current_path != parts[0]+"/": 
                self.folder_data.add(current_path)
        if payload:
        # 2. 判断是时序数据还是关系数据
            if self._is_timeseries(full_topic, payload):
                self._add_to_timeseries(full_topic, payload)
            else:
                self._add_to_relation(full_topic, payload)

    def _is_timeseries(self, topic, payload):
        """
        Determine if it is time series data (rules can be extended)
        :param topic: Contains keywords like elec/fluids or numeric data
        """
        if any(kw in topic for kw in ["electricity", "fluids", "pollution", "production","input","output","chest"]):
            return True
        # # 若 payload 是数值或包含数值字段
        # if isinstance(payload, (int, float)):
        #     return True
        # if isinstance(payload, dict) and any(isinstance(v, (int, float)) for v in payload.values()):
        #     return True
        return False

    def _add_to_relation(self, topic, payload):
        """Add to relation database table"""
        fields = self._extract_fields(payload)
        if fields:  # Only add if fields is not empty
            self.relation_data[topic] = fields

    def _add_to_timeseries(self, topic, payload):
        """Add to time series database table"""
        fields = self._extract_fields(payload)
        if fields:  # Only add if fields is not empty
            self.timeseries_data[topic] = fields

    def _extract_fields(self, payload):
        """
        Extract field names and types from JSON data
        :return: Format like [{"name": "x", "type": "float"}, ...]
        """
        fields = []
        if isinstance(payload, dict):
            for key, value in payload.items():
                field_type = self._infer_type(value)
                fields.append({"name": key, "type": field_type})
        elif isinstance(payload, (list, tuple)) and len(payload) > 0:
            # 处理数组（假设数组元素为同类型）
            sample = payload[0]
            if isinstance(sample, dict):
                for key in sample.keys():
                    field_type = self._infer_type(sample[key])
                    fields.append({"name": key, "type": field_type})
        return fields

    def _infer_type(self, value):
        """Infer field type"""
        if isinstance(value, int):
            return "long" if value > 2147483647 else "int"
        elif isinstance(value, float):
            return "double"
        elif isinstance(value, str):
            return "string"
        elif isinstance(value, bool):
            return "boolean"
        else:
            return "unknown"

    def save_excel(self):
        shutil.copyfile(self.template_filename, self.output_filename)
        wb = load_workbook(self.output_filename)

        # Sheet 1: Folder (Folder hierarchy)
        ws_folder = wb["Folder"]
        start_row = 5
        for idx, path in enumerate(sorted(self.folder_data), start=start_row):
            ws_folder.cell(row=idx, column=1, value=path)

        # Sheet 2: Relation (Relation database)
        ws_relation = wb["relation"]
        for idx, path in enumerate(sorted(self.folder_data), start=start_row):
            ws_relation.cell(row=idx, column=1, value=path)
        
        # Add specific Topic
        for idx, (topic, fields) in enumerate(self.relation_data.items(), start=start_row + len(self.folder_data)):
            if fields:  # Only write if fields is not empty
                ws_relation.cell(row=idx, column=1, value=topic)
                ws_relation.cell(row=idx, column=4, value=json.dumps(fields))
                ws_relation.cell(row=idx, column=6, value="FALSE")  # autoFlow
                ws_relation.cell(row=idx, column=7, value="TRUE")  # autoDashboard
                ws_relation.cell(row=idx, column=8, value="TRUE")  # persistence

        # Sheet 3: TimeSeries (Time series database)
        ws_timeseries = wb["TimeSeries"]
        for idx, path in enumerate(sorted(self.folder_data), start=start_row):
            ws_timeseries.cell(row=idx, column=1, value=path)
        
        # Add specific Topic
        for idx, (topic, fields) in enumerate(self.timeseries_data.items(), start=start_row + len(self.folder_data)):
            if fields:  # Only write if fields is not empty
                ws_timeseries.cell(row=idx, column=1, value=topic)
                ws_timeseries.cell(row=idx, column=4, value=json.dumps(fields))
                ws_timeseries.cell(row=idx, column=6, value="FALSE")  # autoFlow
                ws_timeseries.cell(row=idx, column=7, value="TRUE")  # autoDashboard
                ws_timeseries.cell(row=idx, column=8, value="TRUE")  # persistence (Time series data usually needs persistence)

        wb.save(self.output_filename)
        print(f"Excel generated successfully: {self.output_filename}")

def main():
    template_filename = "template.xlsx"
    output_filename = f"output-{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    excel_gen = ExcelGenerator(template_filename, output_filename)  # Initialize Excel generator
    with open(LOG_FILE, "r") as log_file:
        for row in log_file:
            row = row.strip()  # Remove newline characters
            if ': ' in row:
                topic, msg = row.split(': ', 1)  # Ensure only the first ': ' is split
                excel_gen.add_topic(topic, json.loads(msg))
    excel_gen.save_excel()        

if __name__ == "__main__":
    main()
